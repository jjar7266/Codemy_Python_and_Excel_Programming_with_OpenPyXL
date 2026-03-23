"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Iterate Through Rows of a Spreadsheet
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Iterate Through Rows Using iter_rows()
# ---------------------------------------------------------
for row in ws.iter_rows(  # type: ignore
    min_row=2,
    max_row=10,
    min_col=2,
    max_col=2,
    values_only=True
):
    for cell in row:
        print(cell)

# =========================================================
# Teaching Notes
# =========================================================
"""
1. iter_rows() Overview
   - iter_rows() allows efficient iteration through rows.
   - You can specify row and column boundaries.
   - Setting values_only=True returns raw values instead of Cell objects.

2. Parameter Breakdown
   - min_row / max_row: vertical range
   - min_col / max_col: horizontal range
   - values_only=True: returns Python values (str, int, float, None)

3. Returned Structure
   - Each row is a tuple.
   - Each element in the tuple is a cell value.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Iterate Through Multiple Columns
--------------------------------------------
for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=3, values_only=True):
    print(row)

Example 2 — Iterate Through Entire Sheet
----------------------------------------
for row in ws.iter_rows(values_only=True):
    print(row)

Example 3 — Skip Empty Values
-----------------------------
for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
    for cell in row:
        if cell is not None:
            print(cell)
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting values_only=True and printing Cell objects instead of values.
- Using min_col/max_col incorrectly and selecting the wrong column.
- Expecting iter_rows() to include formatting (it does not).
- Assuming iter_rows() returns a flat list (it returns tuples).
"""

# =========================================================
# Takeaways
# =========================================================
"""
- iter_rows() is the preferred way to loop through structured data.
- values_only=True simplifies output by returning raw values.
- Row iteration is efficient and ideal for reading tables.
- This lesson prepares you for filtering, transforming, and exporting data.
"""
