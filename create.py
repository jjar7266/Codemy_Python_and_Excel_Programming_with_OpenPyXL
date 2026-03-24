"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Create A Spreadsheet Workbook
"""

# ---------------------------------------------------------
# Import Modules
# ---------------------------------------------------------
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Create a New Workbook
# ---------------------------------------------------------
wb = Workbook()

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active

# ---------------------------------------------------------
# Set Worksheet Title
# ---------------------------------------------------------
ws.title = "Names and Colors"  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet
# ---------------------------------------------------------
wb.save("colors.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Workbook()
   - Creates a brand‑new Excel file in memory.
   - Starts with one default worksheet.

2. Active Worksheet
   - wb.active returns the first sheet.
   - You can rename it using ws.title.

3. Saving Files
   - wb.save("filename.xlsx") writes the file to disk.
   - If the file exists, it will be overwritten.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Create Multiple Sheets
----------------------------------
ws1 = wb.create_sheet("Sheet1")
ws2 = wb.create_sheet("Sheet2")

Example 2 — Create a Workbook and Add Data
------------------------------------------
wb = Workbook()
ws = wb.active
ws["A1"] = "Hello"
ws["B1"] = "World"
wb.save("hello_world.xlsx")

Example 3 — Create a Workbook with a Custom First Sheet
-------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "MyData"
wb.save("mydata.xlsx")
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to save the workbook.
- Using load_workbook() when you meant to create a new file.
- Renaming the sheet after saving (must rename before saving).
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Workbook() creates a brand‑new Excel file.
- ws.title lets you rename the first sheet.
- Always save your workbook to write it to disk.
"""
