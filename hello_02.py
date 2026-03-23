"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Grab a Column or Row From a Spreadsheet
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Grab a Whole Column
# ---------------------------------------------------------
column_b = ws["B"]  # type: ignore

# ---------------------------------------------------------
# Loop Through Column Cells
# ---------------------------------------------------------
for cell in column_b:
    print(cell.value)

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Accessing Entire Columns or Rows
   - Columns are accessed using Excel-style letters:
       ws["A"], ws["B"], ws["C"]
   - Rows are accessed using numbers:
       ws[1], ws[2], ws[3]
   - Each returns a tuple of Cell objects.

2. Iterating Through a Column
   - A column like ws["B"] returns all cells in that column.
   - Looping through them allows reading each cell's .value.

3. Cell Objects vs. Values
   - Each item in the column is a Cell object.
   - Use .value to access the stored data.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Grab a Row
----------------------
row_3 = ws[3]
for cell in row_3:
    print(cell.value)

Example 2 — Grab Multiple Columns
---------------------------------
for col in ["A", "B", "C"]:
    for cell in ws[col]:
        print(cell.value)

Example 3 — Skip Empty Cells
----------------------------
for cell in ws["B"]:
    if cell.value is not None:
        print(cell.value)
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to use .value when printing cell contents.
- Using lowercase column letters ("b" instead of "B").
- Assuming ws["B"] returns values instead of Cell objects.
- Attempting to loop through a column that doesn't exist.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Use ws["A"] or ws["B"] to access entire columns.
- Use ws[1] or ws[2] to access entire rows.
- Columns and rows return tuples of Cell objects.
- Looping through them allows processing multiple values.
- This lesson prepares you for reading tables and datasets.
"""
