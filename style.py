"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Change Cell Font, Size, Color, Boldness, Italics
"""

# ---------------------------------------------------------
# Import Modules
# ---------------------------------------------------------
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

# ---------------------------------------------------------
# Create a New Workbook
# ---------------------------------------------------------
wb = Workbook()

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active

# ---------------------------------------------------------
# Select Cells to Style
# ---------------------------------------------------------
cell = ws["A1"]   # type: ignore
cell2 = ws["B1"]  # type: ignore
cell3 = ws["C1"]  # type: ignore

# ---------------------------------------------------------
# Change Font Styles
# ---------------------------------------------------------
cell.font = Font(
    size=30,
    bold=True,
    italic=False,
    color="253bb8"
)

cell2.font = Font(
    size=30,
    bold=False,
    italic=True,
    color="253bb8"
)

cell3.font = Font(
    size=30,
    bold=False,
    italic=False,
    color="253bb8"
)

# ---------------------------------------------------------
# Set Worksheet Title
# ---------------------------------------------------------
ws.title = "Names and Colors"  # type: ignore

# ---------------------------------------------------------
# Create Python Lists of Names, Colors, and Favorite Numbers
# ---------------------------------------------------------
names = ["Dan", "April", "Neal", "Sara"]
colors = ["Blue", "Purple", "Green", "White"]
nums = [12, 39, 42, 21]

# ---------------------------------------------------------
# Add Column Headers
# ---------------------------------------------------------
ws["A1"] = "Names"             # type: ignore
ws["B1"] = "Colors"            # type: ignore
ws["C1"] = "Favorite Numbers"  # type: ignore

# ---------------------------------------------------------
# Adjust Column Widths
# ---------------------------------------------------------
ws.column_dimensions["A"].width = 20  # type: ignore
ws.column_dimensions["B"].width = 20  # type: ignore
ws.column_dimensions["C"].width = 20  # type: ignore

# ---------------------------------------------------------
# Add Names to Worksheet
# ---------------------------------------------------------
starting_name_row = 2
for name in names:
    ws.cell(row=starting_name_row, column=1).value = name  # type: ignore
    starting_name_row += 1

# ---------------------------------------------------------
# Add Colors to Worksheet
# ---------------------------------------------------------
starting_color_row = 2
for color in colors:
    ws.cell(row=starting_color_row, column=2).value = color  # type: ignore
    starting_color_row += 1

# ---------------------------------------------------------
# Add Favorite Numbers to Worksheet
# ---------------------------------------------------------
starting_num_row = 2
for number in nums:
    ws.cell(row=starting_num_row, column=3).value = number  # type: ignore
    starting_num_row += 1

# ---------------------------------------------------------
# Add Excel Formula (SUM)
# ---------------------------------------------------------
ws["C6"] = "=SUM(C2:C5)"  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet
# ---------------------------------------------------------
wb.save("styles.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Font Styling
   - The Font() class controls size, bold, italic, and color.
   - Colors must be hex strings without the '#' symbol.

2. Cell Assignment
   - Styling must be applied BEFORE overwriting the cell value.
   - Setting ws["A1"] after styling replaces the text but keeps the style.

3. Column Widths
   - Excel does not auto-size columns.
   - Setting widths ensures headers do not overflow.

4. Data Entry
   - Lists + loops keep data aligned and easy to maintain.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Underlined Text
---------------------------
ws["A1"].font = Font(size=20, underline="single")

Example 2 — Different Font Family
---------------------------------
ws["B1"].font = Font(name="Calibri", size=24, bold=True)

Example 3 — Color by Hex Code
-----------------------------
ws["C1"].font = Font(color="FF0000")  # Red
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to apply Font() before overwriting the cell value.
- Using '#' in hex colors (OpenPyXL requires plain hex).
- Expecting Excel to auto-size columns (it won't).
- Using lowercase column letters in formulas.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Font styling in OpenPyXL is controlled through the Font() class.
- You can change size, bold, italic, underline, and color.
- Styling must be applied before saving the workbook.
- Combining styling + data entry creates professional spreadsheets.
"""
