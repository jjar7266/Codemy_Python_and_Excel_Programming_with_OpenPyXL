"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Create a Line Chart in a Spreadsheet with Python
"""

# ---------------------------------------------------------
# Import Modules
# ---------------------------------------------------------
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart, LineChart3D, Reference
)

# ---------------------------------------------------------
# Create a New Workbook
# ---------------------------------------------------------
wb = Workbook()

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active

# ---------------------------------------------------------
# Set Worksheet Title
# ---------------------------------------------------------
ws.title = "New Worksheet"  # type: ignore

# ---------------------------------------------------------
# Create Python Lists of Names, Colors, and Salary Data
# ---------------------------------------------------------
names = ["John", "Erin", "Sam", "Tina", "Josh", "Mary", "Bob", "Lisa", "Steve"]
colors = ["Blue", "Red", "Pink", "Green", "Yellow", "Black", "White", "Purple", "Gray"]
nums = [180000, 190000, 120000, 89000, 42000, 12000, 11800, 79000, 210000]

# ---------------------------------------------------------
# Add Column Headers
# ---------------------------------------------------------
ws["A1"] = "Names"             # type: ignore
ws["B1"] = "Colors"            # type: ignore
ws["C1"] = "Salary"            # type: ignore

# ---------------------------------------------------------
# Adjust Column Widths
# ---------------------------------------------------------
ws.column_dimensions["A"].width = 12  # type: ignore
ws.column_dimensions["B"].width = 12  # type: ignore
ws.column_dimensions["C"].width = 16  # type: ignore

# ---------------------------------------------------------
# Add Names to Worksheet
# ---------------------------------------------------------
starting_name_row = 2
for name in names:
    ws.cell(row=starting_name_row, column=1).value = name  # type: ignore
    starting_name_row += 1

# ---------------------------------------------------------
# Add Colors to Worksheet
# ---------------------------------------------------------
starting_color_row = 2
for color in colors:
    ws.cell(row=starting_color_row, column=2).value = color  # type: ignore
    starting_color_row += 1

# ---------------------------------------------------------
# Add Salary Data to Worksheet
# ---------------------------------------------------------
starting_num_row = 2
for number in nums:
    ws.cell(row=starting_num_row, column=3).value = number  # type: ignore
    starting_num_row += 1

# ---------------------------------------------------------
# Create Line Chart
# ---------------------------------------------------------
chart = LineChart()   # Use LineChart3D() for 3D version

# Labels (Names)
labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=10)  # type: ignore

# Data (Salary)
data = Reference(ws, min_col=3, min_row=1, max_row=10)               # type: ignore

# Add data to chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = "Employee Salaries"

# ---------------------------------------------------------
# Place the Chart on the Worksheet
# ---------------------------------------------------------
ws.add_chart(chart, "E2")  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet
# ---------------------------------------------------------
wb.save("hello_04.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. LineChart vs LineChart3D
   - LineChart() creates a standard 2D line chart.
   - LineChart3D() creates a 3D version.
   - Both use the same Reference() structure.

2. Data Setup
   - Column A = Names (labels)
   - Column C = Salary (values)
   - Reference() defines the exact cell ranges.

3. Chart Behavior
   - Line charts are great for trends over time.
   - Even though this data isn't time‑based, the chart still works.

4. Chart Placement
   - ws.add_chart(chart, "E2") places the chart starting at E2.
   - Excel renders the chart when the file is opened.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — 3D Line Chart
-------------------------
chart = LineChart3D()

Example 2 — Smooth Line
-----------------------
chart.smooth = True

Example 3 — Add Axis Titles
---------------------------
chart.x_axis.title = "Employees"
chart.y_axis.title = "Salary"

Example 4 — Change Chart Style
------------------------------
chart.style = 13
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Using lowercase column letters in Reference().
- Forgetting titles_from_data=True.
- Creating the chart before writing the data.
- Not adding the chart to the worksheet.
- Saving the file before adding the chart.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Line charts are created using LineChart() or LineChart3D().
- Reference() defines the data and label ranges.
- Charts must be added to the worksheet manually.
- Excel renders the chart when the file is opened.
- This completes the core chart types in OpenPyXL.
"""
