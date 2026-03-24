"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Add Borders To Cells in a Spreadsheet
"""

# ---------------------------------------------------------
# Import Modules
# ---------------------------------------------------------
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# ---------------------------------------------------------
# Create a New Workbook
# ---------------------------------------------------------
wb = Workbook()

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active

# ---------------------------------------------------------
# Select Cells to Style
# ---------------------------------------------------------
cell = ws["A1"]   # type: ignore
cell2 = ws["B1"]  # type: ignore
cell3 = ws["C1"]  # type: ignore

# ---------------------------------------------------------
# Change Font Styles
# ---------------------------------------------------------
cell.font = Font(size=30, bold=True, italic=False, color="253bb8")
cell2.font = Font(size=30, bold=False, italic=True, color="253bb8")
cell3.font = Font(size=30, bold=False, italic=False, color="253bb8")

# ---------------------------------------------------------
# Define Border Style
# ---------------------------------------------------------
my_bd = Side(style="double", color="000000")

# Example border on a single cell
B3 = ws["B3"]  # type: ignore
B3.border = Border(left=my_bd, right=my_bd, top=my_bd, bottom=my_bd)

# Add bottom borders to header cells
cell.border = Border(bottom=my_bd)
cell2.border = Border(bottom=my_bd)
cell3.border = Border(bottom=my_bd)

# ---------------------------------------------------------
# Set Worksheet Title
# ---------------------------------------------------------
ws.title = "Names and Colors"  # type: ignore

# ---------------------------------------------------------
# Create Python Lists of Names, Colors, and Favorite Numbers
# ---------------------------------------------------------
names = ["Dan", "April", "Neal", "Sara"]
colors = ["Blue", "Purple", "Green", "White"]
nums = [12, 39, 42, 21]

# ---------------------------------------------------------
# Add Column Headers
# ---------------------------------------------------------
ws["A1"] = "Names"             # type: ignore
ws["B1"] = "Colors"            # type: ignore
ws["C1"] = "Favorite Numbers"  # type: ignore

# ---------------------------------------------------------
# Adjust Column Widths
# ---------------------------------------------------------
ws.column_dimensions["A"].width = 20  # type: ignore
ws.column_dimensions["B"].width = 20  # type: ignore
ws.column_dimensions["C"].width = 20  # type: ignore

# ---------------------------------------------------------
# Add Names to Worksheet
# ---------------------------------------------------------
starting_name_row = 2
for name in names:
    ws.cell(row=starting_name_row, column=1).value = name  # type: ignore
    starting_name_row += 1

# ---------------------------------------------------------
# Add Colors to Worksheet
# ---------------------------------------------------------
starting_color_row = 2
for color in colors:
    ws.cell(row=starting_color_row, column=2).value = color  # type: ignore
    starting_color_row += 1

# ---------------------------------------------------------
# Add Favorite Numbers to Worksheet
# ---------------------------------------------------------
starting_num_row = 2
for number in nums:
    ws.cell(row=starting_num_row, column=3).value = number  # type: ignore
    starting_num_row += 1

# ---------------------------------------------------------
# Add Excel Formula (SUM)
# ---------------------------------------------------------
ws["C6"] = "=SUM(C2:C5)"  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet
# ---------------------------------------------------------
wb.save("styles.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Borders in OpenPyXL
   - Borders are created using Border() and Side().
   - Side() controls the line style and color.
   - Common styles: thin, medium, thick, double, dashed.

2. Applying Borders
   - Borders must be applied to each side individually.
   - Border(bottom=my_bd) applies only the bottom border.

3. Styling Order
   - You can style a cell before or after assigning its value.
   - Borders and fonts persist even if the value changes.

4. Practical Use
   - Borders help separate headers, sections, and totals.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Thin Border Around a Cell
-------------------------------------
thin = Side(style="thin", color="000000")
ws["A5"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

Example 2 — Thick Bottom Border Only
------------------------------------
thick = Side(style="thick", color="000000")
ws["A1"].border = Border(bottom=thick)

Example 3 — Dashed Border
-------------------------
dash = Side(style="dashed", color="FF0000")
ws["B2"].border = Border(left=dash, right=dash)
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to apply borders to all four sides.
- Using invalid style names (must match OpenPyXL docs).
- Expecting Excel to auto-size columns (it won't).
- Overwriting a styled cell without reapplying formatting.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Borders are created using Border() + Side().
- You can style any side of a cell independently.
- Borders help organize and visually separate spreadsheet data.
- Combining borders, fonts, and formulas creates professional Excel files.
"""
