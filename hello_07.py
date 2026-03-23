"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Add Cells to a Spreadsheet (Second Method)
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Change Many Cells Using ws.cell()
# ---------------------------------------------------------
starting_row = 11

ws.cell(row=starting_row, column=1).value = "Neo"    # type: ignore
ws.cell(row=starting_row, column=2).value = "Black"  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet as a New File
# ---------------------------------------------------------
wb.save("hello_2.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Using ws.cell()
   - ws.cell(row=?, column=?) gives direct numeric access.
   - Useful when working with loops or dynamic positions.
   - Equivalent to ws["A1"], but more flexible.

2. Assigning Values
   - .value is used to set the cell’s content.
   - Accepts strings, numbers, booleans, or formulas.

3. Why Use ws.cell()?
   - Perfect for programmatic updates.
   - Ideal when row/column numbers come from variables.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Fill a Row Dynamically
----------------------------------
for col, value in enumerate(["Alice", "Blue", 42], start=1):
    ws.cell(row=12, column=col).value = value

Example 2 — Fill Multiple Rows
------------------------------
for r in range(13, 18):
    ws.cell(row=r, column=1).value = f"Name {r}"
    ws.cell(row=r, column=2).value = "Color"

Example 3 — Insert Numbers
--------------------------
for i in range(1, 6):
    ws.cell(row=20, column=i).value = i * 10
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting .value when assigning data.
- Using row/column indexes starting at 0 (Excel starts at 1).
- Overwriting existing data unintentionally.
- Saving to the same file without realizing it overwrites.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- ws.cell() is the numeric, programmatic way to access cells.
- Perfect for loops and dynamic updates.
- Always save after modifying the workbook.
- This method prepares you for building automated Excel writers.
"""
