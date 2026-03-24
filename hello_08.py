"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Loop Through a Spreadsheet and Add Names
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Create Python List of Names
# ---------------------------------------------------------
names = ["Dan", "April", "Neal"]

# ---------------------------------------------------------
# Loop Through List and Add Names to Spreadsheet
# ---------------------------------------------------------
starting_row = 12

for name in names:
    ws.cell(row=starting_row, column=1).value = name   # type: ignore
    starting_row += 1

# ---------------------------------------------------------
# Save Spreadsheet as a New File
# ---------------------------------------------------------
wb.save("hello_2.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Looping Through Lists
   - A Python list is perfect for batch updates.
   - Each loop iteration writes one name into the sheet.

2. Dynamic Row Movement
   - starting_row tracks where to write.
   - Incrementing it moves the cursor down each loop.

3. ws.cell() in Loops
   - Ideal for programmatic updates.
   - Lets you write data row-by-row or column-by-column.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Add Names and Ages
------------------------------
people = [("Dan", 30), ("April", 28), ("Neal", 35)]
row = 12
for name, age in people:
    ws.cell(row=row, column=1).value = name
    ws.cell(row=row, column=2).value = age
    row += 1

Example 2 — Add Names to a Different Column
-------------------------------------------
row = 12
for name in names:
    ws.cell(row=row, column=3).value = name
    row += 1

Example 3 — Auto-Number Rows
----------------------------
row = 12
for i, name in enumerate(names, start=1):
    ws.cell(row=row, column=1).value = i
    ws.cell(row=row, column=2).value = name
    row += 1
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to increment the row number inside the loop.
- Starting rows at 0 (Excel rows begin at 1).
- Overwriting existing data unintentionally.
- Saving to the same file and losing original data.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Loops + ws.cell() = powerful automated data entry.
- Lists make it easy to batch-insert rows.
- Always track your row index when writing sequentially.
- Save your workbook after making changes.
"""
