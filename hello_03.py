"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Grab Ranges of Cells
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Grab a Range of Cells
# ---------------------------------------------------------

# returns a tuple of rows, each row is a tuple of Cell objects
cell_range = ws["A2":"B10"]  # type: ignore


# ---------------------------------------------------------
# Loop Through the Range
# ---------------------------------------------------------
for row in cell_range:
    for cell in row:
        print(cell.value)

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Accessing a Range of Cells
   - Use Excel-style slicing:
       ws["A2":"B10"]
   - This returns a tuple of rows.
   - Each row is a tuple of Cell objects.

2. Nested Loop Structure
   - Outer loop: iterates through each row in the range.
   - Inner loop: iterates through each cell in that row.
   - Use .value to access the stored data.

3. Range Behavior
   - Ranges preserve the grid structure.
   - Useful for reading tables, blocks, and datasets.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Print Only One Column From the Range
------------------------------------------------
for row in ws["A2":"B10"]:
    print(row[0].value)   # first column in the range

Example 2 — Convert Range to a List of Values
---------------------------------------------
values = [[cell.value for cell in row] for row in ws["A2":"B10"]]
print(values)

Example 3 — Skip Empty Cells
----------------------------
for row in ws["A2":"B10"]:
    for cell in row:
        if cell.value is not None:
            print(cell.value)
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Using parentheses instead of brackets for ranges.
- Forgetting that ranges return tuples of rows, not a flat list.
- Attempting to access .value directly on the range.
- Expecting ws["A2":"B10"] to include formatting (it does not).
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Use ws["A2":"B10"] to access rectangular blocks of cells.
- Ranges return rows, and each row contains Cell objects.
- Nested loops allow processing each cell in the block.
- This lesson prepares you for reading tables and structured data.
"""
