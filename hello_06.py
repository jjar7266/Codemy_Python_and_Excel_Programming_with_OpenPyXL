"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Change Cells and Save Spreadsheet
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Change a Cell Value
# ---------------------------------------------------------
ws["A2"] = "Johnny"  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet as a New File
# ---------------------------------------------------------
wb.save("hello_2.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Changing Cell Values
   - Assigning a value is as simple as:
       ws["A2"] = "New Value"
   - You can assign strings, numbers, booleans, or formulas.

2. Saving the Workbook
   - wb.save("filename.xlsx") writes the file to disk.
   - If the file already exists, it will be overwritten.
   - Saving to a new filename preserves the original.

3. Active Worksheet
   - ws = wb.active selects the first sheet by default.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Change Multiple Cells
---------------------------------
ws["A3"] = "Hello"
ws["B3"] = 123
ws["C3"] = "=SUM(B1:B10)"

Example 2 — Change Cells in a Loop
----------------------------------
for row in range(2, 6):
    ws[f"B{row}"] = row * 10

Example 3 — Save Over Existing File
-----------------------------------
wb.save("hello.xlsx")   # overwrites original
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to save after making changes.
- Saving to the same filename unintentionally overwrites data.
- Using lowercase column letters ("a2" instead of "A2").
- Expecting formatting to be preserved when only values are changed.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Changing a cell is as simple as assigning a value.
- Always save after modifying the workbook.
- Saving to a new filename preserves your original data.
- This lesson prepares you for writing dynamic data into spreadsheets.
"""
