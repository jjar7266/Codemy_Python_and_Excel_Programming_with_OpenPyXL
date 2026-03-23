"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Grab A Cell From A Spreadsheet
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Read Specific Cells
# ---------------------------------------------------------
name = ws["A3"].value  # type: ignore
color = ws["B3"].value  # type: ignore

# ---------------------------------------------------------
# Output Result
# ---------------------------------------------------------
print(f"{name}: {color}")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Accessing a Worksheet
   - wb.active returns the currently active worksheet.
   - This is typically the first sheet unless the workbook specifies otherwise.

2. Accessing a Cell
   - Use Excel-style coordinates such as "A3" or "B3".
   - ws["A3"] returns a Cell object.
   - ws["A3"].value returns the actual stored value.

3. Storing Cell Values
   - Assigning cell values to variables improves readability and organization.
   - Example:
       name = ws["A3"].value
       color = ws["B3"].value

4. Printing Formatted Output
   - f-strings allow clean formatting:
       print(f"{name}: {color}")
   - Example output:
       John: Blue
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Accessing Multiple Rows
-----------------------------------
first = ws["A2"].value
second = ws["A3"].value
third = ws["A4"].value
print(first, second, third)

Example 2 — Looping Through a Column
------------------------------------
for row in range(2, 6):
    print(ws[f"A{row}"].value)

Example 3 — Checking for Empty Cells
------------------------------------
value = ws["C3"].value
if value is None:
    print("Cell is empty")
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting to use .value and printing the Cell object instead.
- Using lowercase column letters ("a3" instead of "A3").
- Attempting to access a cell outside the sheet’s used range.
- Assuming wb.active can return None (it does not in OpenPyXL).
"""

# =========================================================
# Takeaways
# =========================================================
"""
- Use Excel-style coordinates to access cells.
- Always use .value to retrieve the stored data.
- Assigning values to variables improves clarity.
- f-strings provide clean, readable output.
- This lesson forms the foundation for reading, looping, and modifying spreadsheets.
"""
