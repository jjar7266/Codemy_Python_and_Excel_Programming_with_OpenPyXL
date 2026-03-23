"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Iterate Through Columns of a Spreadsheet
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# ---------------------------------------------------------
# Load Existing Spreadsheet
# ---------------------------------------------------------
wb = load_workbook("hello.xlsx")

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active  # type: ignore

# ---------------------------------------------------------
# Iterate Through Columns Using iter_cols()
# ---------------------------------------------------------
for col in ws.iter_cols(  # type: ignore
    min_row=1,
    max_row=10,
    min_col=2,
    max_col=2,
    values_only=True
):
    for cell in col:
        print(cell)

# =========================================================
# Teaching Notes
# =========================================================
"""
1. iter_cols() Overview
   - iter_cols() works like iter_rows(), but iterates vertically.
   - Useful for reading down a column or multiple columns.
   - Setting values_only=True returns raw Python values.

2. Parameter Breakdown
   - min_row / max_row: vertical boundaries
   - min_col / max_col: horizontal boundaries
   - values_only=True: returns values instead of Cell objects

3. Returned Structure
   - Each column is a tuple.
   - Each element in the tuple is a cell value.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Iterate Through Multiple Columns
--------------------------------------------
for col in ws.iter_cols(min_row=1, max_row=10, min_col=1, max_col=3, values_only=True):
    print(col)

Example 2 — Iterate Through Entire Sheet by Columns
---------------------------------------------------
for col in ws.iter_cols(values_only=True):
    print(col)

Example 3 — Skip Empty Values
-----------------------------
for col in ws.iter_cols(min_row=1, max_row=10, values_only=True):
    for cell in col:
        if cell is not None:
            print(cell)
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting values_only=True and printing Cell objects.
- Mixing up iter_rows() and iter_cols().
- Using incorrect min_col/max_col values.
- Expecting formatting or formulas (iterators return values only).
"""

# =========================================================
# Takeaways
# =========================================================
"""
- iter_cols() is ideal for vertical data processing.
- values_only=True simplifies output.
- Columns return tuples of values.
- This lesson prepares you for scanning, filtering, and analyzing column data.
"""
