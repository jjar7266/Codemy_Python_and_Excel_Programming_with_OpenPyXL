"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Create a Pie Chart in a Spreadsheet with Python
"""

# ---------------------------------------------------------
# Import Modules
# ---------------------------------------------------------
from openpyxl import Workbook
from openpyxl.chart import PieChart, PieChart3D, Reference

# ---------------------------------------------------------
# Create a New Workbook
# ---------------------------------------------------------
wb = Workbook()

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active

# ---------------------------------------------------------
# Set Worksheet Title
# ---------------------------------------------------------
ws.title = "New Worksheet"  # type: ignore

# ---------------------------------------------------------
# Create Python Lists of Names, Colors, and Salary Data
# ---------------------------------------------------------
names = ["John", "Erin", "Sam", "Tina", "Josh", "Mary", "Bob", "Lisa", "Steve"]
colors = ["Blue", "Red", "Pink", "Green", "Yellow", "Black", "White", "Purple", "Gray"]
nums = [180000, 190000, 120000, 89000, 42000, 12000, 11800, 79000, 210000]

# ---------------------------------------------------------
# Add Column Headers
# ---------------------------------------------------------
ws["A1"] = "Names"             # type: ignore
ws["B1"] = "Colors"            # type: ignore
ws["C1"] = "Salary"            # type: ignore

# ---------------------------------------------------------
# Adjust Column Widths
# ---------------------------------------------------------
ws.column_dimensions["A"].width = 12  # type: ignore
ws.column_dimensions["B"].width = 12  # type: ignore
ws.column_dimensions["C"].width = 16  # type: ignore

# ---------------------------------------------------------
# Add Names to Worksheet
# ---------------------------------------------------------
starting_name_row = 2
for name in names:
    ws.cell(row=starting_name_row, column=1).value = name  # type: ignore
    starting_name_row += 1

# ---------------------------------------------------------
# Add Colors to Worksheet
# ---------------------------------------------------------
starting_color_row = 2
for color in colors:
    ws.cell(row=starting_color_row, column=2).value = color  # type: ignore
    starting_color_row += 1

# ---------------------------------------------------------
# Add Salary Data to Worksheet
# ---------------------------------------------------------
starting_num_row = 2
for number in nums:
    ws.cell(row=starting_num_row, column=3).value = number  # type: ignore
    starting_num_row += 1

# ---------------------------------------------------------
# Create Pie Chart
# ---------------------------------------------------------
chart = PieChart()

# Labels (Names)
labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=10)  # type: ignore

# Data (Salary)
data = Reference(ws, min_col=3, min_row=1, max_row=10)               # type: ignore

# Add data to chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = "Employee Salaries"

# ---------------------------------------------------------
# Place the Chart on the Worksheet
# ---------------------------------------------------------
ws.add_chart(chart, "E2")        # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet
# ---------------------------------------------------------
wb.save("hello_02.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Workbook and Worksheet
   - Workbook() creates a brand‑new Excel file.
   - wb.active selects the first worksheet.
   - ws.title renames the sheet.

2. Data Setup
   - Lists store names, colors, and salary values.
   - Loops write each list into its own column.

3. Pie Chart Basics
   - PieChart() creates a 2D pie chart.
   - Reference() defines the data range for labels and values.
   - chart.add_data(..., titles_from_data=True) tells Excel to use the header.
   - chart.set_categories(labels) assigns the labels.

4. Chart Placement
   - ws.add_chart(chart, "E2") places the chart starting at cell E2.
   - Excel renders the chart when the file is opened.

5. Important
   - Charts do NOT render in Python — only in Excel.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — 3D Pie Chart
------------------------
chart = PieChart3D()
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = "3D Salary Chart"
ws.add_chart(chart, "E2")

Example 2 — Change Chart Style
------------------------------
chart.style = 10

Example 3 — Add Data Labels
---------------------------
chart.dataLabels = chart.dataLabels or DataLabelList()
chart.dataLabels.showVal = True
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Creating the chart before writing the data.
- Using lowercase column letters in Reference().
- Forgetting titles_from_data=True.
- Not adding the chart to the worksheet.
- Saving the file before adding the chart.
"""

# =========================================================
# Takeaways
# =========================================================
"""
- OpenPyXL can generate Excel charts programmatically.
- Pie charts require both labels and values.
- Excel handles the rendering of the chart.
- You can place charts anywhere on the sheet.
- This is the foundation for bar charts, line charts, and more.
"""
