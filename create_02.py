"""
Codemy.com
Course: Python and Excel Programming With OpenPyXL
Instructor: John Elder
Student: Jose "Joe" Ruiz

Lesson: Use Excel Formulas With Python
"""

# ---------------------------------------------------------
# Import Modules
# ---------------------------------------------------------
from openpyxl import Workbook
from openpyxl import load_workbook  # Included for course continuity

# ---------------------------------------------------------
# Create a New Workbook
# ---------------------------------------------------------
wb = Workbook()

# ---------------------------------------------------------
# Select Active Worksheet
# ---------------------------------------------------------
ws = wb.active

# ---------------------------------------------------------
# Set Worksheet Title
# ---------------------------------------------------------
ws.title = "Names and Colors"  # type: ignore

# ---------------------------------------------------------
# Create Python Lists of Names, Colors, and Favorite Numbers
# ---------------------------------------------------------
names = ["Dan", "April", "Neal", "Sara"]
colors = ["Blue", "Purple", "Green", "White"]
nums = [12, 39, 42, 21]

# ---------------------------------------------------------
# Add Column Headers
# ---------------------------------------------------------
ws["A1"] = "Names"              # type: ignore
ws["B1"] = "Colors"             # type: ignore
ws["C1"] = "Favorite Numbers"   # type: ignore

# ---------------------------------------------------------
# Adjust Column Widths
# ---------------------------------------------------------
ws.column_dimensions["A"].width = 10  # type: ignore
ws.column_dimensions["B"].width = 10  # type: ignore
ws.column_dimensions["C"].width = 16  # type: ignore

# ---------------------------------------------------------
# Add Names to Worksheet
# ---------------------------------------------------------
starting_name_row = 2
for name in names:
    ws.cell(row=starting_name_row, column=1).value = name  # type: ignore
    starting_name_row += 1

# ---------------------------------------------------------
# Add Colors to Worksheet
# ---------------------------------------------------------
starting_color_row = 2
for color in colors:
    ws.cell(row=starting_color_row, column=2).value = color  # type: ignore
    starting_color_row += 1

# ---------------------------------------------------------
# Add Favorite Numbers to Worksheet
# ---------------------------------------------------------
starting_num_row = 2
for number in nums:
    ws.cell(row=starting_num_row, column=3).value = number  # type: ignore
    starting_num_row += 1

# ---------------------------------------------------------
# Add Excel Formula (SUM)
# ---------------------------------------------------------
ws["C6"] = "=SUM(C2:C5)"  # type: ignore

# ---------------------------------------------------------
# Save Spreadsheet
# ---------------------------------------------------------
wb.save("colors.xlsx")

print("File Was Saved...")

# =========================================================
# Teaching Notes
# =========================================================
"""
1. Excel Formulas in Python
   - OpenPyXL allows you to write formulas exactly as you would in Excel.
   - The formula is stored in the cell, and Excel evaluates it when opened.

2. SUM Formula
   - "=SUM(C2:C5)" adds all numbers in the range.
   - The formula is written as a string.

3. Column Widths
   - Excel does not auto-size columns.
   - Setting widths prevents header text from overflowing into the next column.

4. Data Alignment
   - Each list is written into its own column.
   - Row counters ensure data stays aligned.
"""

# =========================================================
# Example Variations
# =========================================================
"""
Example 1 — Dynamic Formula Row
-------------------------------
last_row = len(nums) + 1
ws[f"C{last_row + 1}"] = f"=SUM(C2:C{last_row})"

Example 2 — Average Formula
---------------------------
ws["D1"] = "Average"
ws["D6"] = "=AVERAGE(C2:C5)"

Example 3 — Count Formula
-------------------------
ws["E1"] = "Count"
ws["E6"] = "=COUNT(C2:C5)"
"""

# =========================================================
# Common Mistakes
# =========================================================
"""
- Forgetting the '=' sign when writing formulas.
- Using lowercase column letters (Excel requires uppercase).
- Saving to the wrong filename and overwriting previous lessons.
- Expecting Python to evaluate the formula (Excel evaluates it, not Python).
"""

# =========================================================
# Takeaways
# =========================================================
"""
- OpenPyXL can write Excel formulas directly into cells.
- Excel evaluates formulas when the file is opened.
- Column widths help keep the sheet readable.
- Lists + loops + formulas = powerful automated spreadsheets.
"""
